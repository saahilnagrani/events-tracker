from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date

B = "https://dubai.platinumlist.net/event-tickets/"
A = "https://abu-dhabi.platinumlist.net/event-tickets/"

# city, category, event, artist, start, end, time, venue, price, language, notes, url
rows = [
("Dubai","Comedy","Cocktails & Comedy Show","Dubai stand-up line-up (not named)",date(2026,8,12),None,"20:00","Bondz Bar & Lounge, Emirates Financial Towers, DIFC",85,"English","",B+"cocktails-comedy-show"),
("Dubai","Comedy",'The Laughter Factory Premium Comedy Club "August 2026 Tour"',"Ricky Balshaw, Tevin Everett, Abz Ali, Miqdaad Dohadwala, Andrea Brooks",date(2026,8,13),date(2026,8,15),"20:00-20:30","Multi-venue: Bla Bla (JBR), Radisson DAMAC Hills, Dukes The Palm",175,"English","Tour also played Abu Dhabi (The Club, Mina Zayed) on 9 Aug",'https://platinumlist.net/event-tickets/107168/the-laughter-factory-premium-comedy-club-august-2026-tour'),
("Dubai","Desi","Chitthi Aayi Hai - The Pitaara Festival","Ensemble (bhajan, qawwali, ghazal, Sufi, haveli sangeet)",date(2026,8,14),date(2026,8,16),"Not stated","The Junction, Alserkal Avenue, Al Quoz",45,"Hindi, English, Urdu & Marwari","",B+"107593/chitthi-aayi-hai-the-pitaara-festival-the-festival-that-calls-you-home-in-dubai"),
("Dubai","Desi","Qawalli Night Live at Troy","Folk of Khan",date(2026,8,14),None,"20:00","Troy, Ramee Dream Hotel, Business Bay",75,"Not stated","",B+"107578/qawalli-night-live-at-troy-in-dubai"),
("Dubai","Desi","Sufi Mehfil at Troy","Not named (Sufi vocals, harmonium, dholak, khartal)",date(2026,8,15),None,"20:00","Troy, Ramee Dream Hotel, Business Bay",75,"Not stated","",B+"107542/sufi-mehfil-at-troy-in-dubai"),
("Dubai","Comedy","Joke Hub - Madcat Comedy Show","Rotating Mad Cat Comedy Club line-up",date(2026,8,15),date(2026,8,29),"20:30-22:30","Skafos, Canopy by Hilton Dubai Al Seef, Bur Dubai",85,"English","Runs every Saturday",B+"joke-hub-madcat-comedy-show"),
("Dubai","Comedy","The Social Brunch","No performer listed (live DJ)",date(2026,8,15),date(2026,8,29),"13:00-16:00","Skafos, Canopy by Hilton Dubai Al Seef, Bur Dubai",146.25,"English","Comedy brunch; flash-sale price, standard AED 195",B+"the-social-brunch"),
("Dubai","Desi","Lucky Ali in Dubai","Lucky Ali",date(2026,8,16),None,"20:30","Coca-Cola Arena, City Walk",175,"Not stated","",B+"107004/lucky-ali"),
("Dubai","Desi","Baithak Live at Troy","Not named (Indian folk)",date(2026,8,16),None,"20:00","Troy, Ramee Dream Hotel, Business Bay",75,"Not stated","",B+"107543/baithak-live-at-troy-in-dubai"),
("Dubai","Comedy","Wildest Standup Comedy by Mad Cat","English stand-up line-up (not named)",date(2026,8,16),None,"20:00","Headlines Premier by Citrus, Dubai Marina",99,"English","",B+"wildest-standup-comedy-by-mad-cat"),
("Dubai","Comedy","The Coterie - The Comedy Lounge","Rotating line-up of 5-6 UAE comedians",date(2026,8,19),None,"20:30-22:30","Coterie Social, Ibn Battuta",85,"English","Every other Wednesday",B+"the-coterie-the-comedy-lounge"),
("Dubai","Desi","Sunil Grover in Dubai","Sunil Grover",date(2026,8,21),None,"20:00","Coca-Cola Arena, City Walk",125,"Hindi","",B+"107289/sunil-grover-in-dubai"),
("Dubai","Comedy","Comedy @ the Speakeasy","Rotating local comedians",date(2026,8,23),None,"19:00-21:30","Moon Bar by Sana, Jumeirah Mina A'Salam, Madinat Jumeirah",85,"Not stated","",B+"the-speakeasy-madcat-comedy-show"),
("Dubai","Comedy","Rami Jbr Stand Up Comedy","Rami Jbr",date(2026,8,28),None,"20:00","JAFZA One Convention Centre, Jebel Ali",150,"Arabic","",B+"107007/rami-jbr-stand-up-comedy-in-dubai"),
("Dubai","Comedy","Jimmy Carr 'Laughs Funny' Live at Dubai Opera","Jimmy Carr",date(2026,8,29),date(2026,8,30),"18:30 & 21:30","Dubai Opera, Downtown Dubai",250,"English","Two shows per night",B+"107320/jimmy-carr-laughs-funny-live-at-dubai-opera"),
("Dubai","Comedy + Desi","Gaurav Kapoor Live in Dubai","Gaurav Kapoor",date(2026,8,29),None,"20:30","Dubai World Trade Centre, Sheikh Zayed Road",125,"Not stated","",B+"107353/gaurav-kapoor-live-in-dubai"),
("Dubai","Comedy","Mohamed Abdelaty Live In Dubai, Sa3a Mally","Mohamed Abdelaty",date(2026,8,29),None,"19:30","The New Covent Garden Theatre, Mall of the Emirates",245,"Arabic","",B+"107527/mohamed-abdelaty-live-in-dubai-sa3a-mally"),
("Dubai","Comedy","Ahmed & Ahmed - Stand Up Comedy Show","Ahmed Magdy & Ahmed Hassan",date(2026,8,29),None,"21:00","JAFZA One Convention Centre, Jebel Ali",195,"Arabic","",B+"107145/ahmed-ahmed-stand-up-comedy-show-at-jafza-one-convention-centre"),
("Dubai","Comedy","Atul Khatri Live in Dubai","Atul Khatri",date(2026,8,29),None,"20:00","Emirates Theatre, Emirates International School, Jumeirah",95,"Hindi","",B+"105366/atul-khatri-live-in-dubai"),
("Dubai","Desi","Bhuwin Experience - Noor-e-Daastan","Bhuwin (Bhuwin Khursija)",date(2026,8,30),None,"19:00","The New Covent Garden Theatre, Mall of the Emirates",125,"Not stated","",B+"105332/bhuwin-experience-noor-e-daastan"),
("Dubai","Comedy","Mina Nader at Dubai Opera","Mina Nader",date(2026,9,5),None,"22:00","Dubai Opera, Downtown Dubai",195,"Arabic","",B+"107278/mina-nader-at-dubai-opera"),
("Dubai","Comedy","Taha Desouky - Stand Up Comedy Show","Taha Desouky",date(2026,9,6),None,"21:00","Dubai Opera, Downtown Dubai",250,"Arabic","",B+"107253/taha-desouky-stand-up-comedy-show-in-dubai"),
("Dubai","Comedy + Desi","Comedy Box Office by Rajat Chauhan, Siddhartha Shetty & Anmol Garg","Rajat Chauhan, Siddhartha Shetty, Anmol Garg",date(2026,9,12),None,"19:00","Emirates Theatre, Emirates International School, Jumeirah",125,"Not stated","",B+"107404/comedy-box-office-by-rajat-chauhan-and-siddhartha-shetty-and-anmol-garg"),
("Dubai","Comedy","Magic Phil in Mother Goose","Magic Phil",date(2026,9,19),None,"11:00 & 15:00","Theatre by QE2, Queen Elizabeth 2, Port Rashid",79,"English","",B+"107616/magic-phil-in-mother-goose-at-theatre-by-qe2"),
("Dubai","Comedy + Desi","O Sakhi: Her Voice in Verse and Song","Katha Sutra",date(2026,9,19),None,"17:00","Live@Play, Warehouse B17, ABA Avenue, Al Quoz",125,"Multiple","",B+"107805/o-sakhi-her-voice-in-verse-and-song"),
("Dubai","Desi","Radhika Das Lightfall Live","Radhika Das",date(2026,9,20),None,"06:30 as printed","Coca-Cola Arena, City Walk",150,"Not stated","Event page prints 06:30 with no AM/PM - confirm time before booking",B+"107296/radhika-das-lightfall"),
("Dubai","Comedy + Desi","Phulka Dots - A Musical Stand-Up Comedy Show","MOMic Amruta (Amruta Bendre)",date(2026,9,20),None,"17:00","Live@Play, Warehouse B17, ABA Avenue, Al Quoz",100,"Hinglish","",B+"107415/phulka-dots-a-musical-stand-up-comedy-show-ft-momic-amruta"),
("Dubai","Desi","Bhakti 2.0","O.S. Arun, Vinay Varanasi, Vishaka Hari",date(2026,9,25),date(2026,9,27),"18:00","Zabeel Ladies Club (listing title says Oud Mehta Theater)",80,"English & Tamil","Early-bird price to 31 Aug; standard AED 100",B+"107427/bhakti-at-oud-mehta-theater-in-dubai"),
("Dubai","Comedy","3a Ka3ba Improv Comedy Show","Fouad Yammine, Anthony Hamawi, Tony Dagher, Jad Bou Karam",date(2026,9,25),None,"21:00","JAFZA One Convention Centre, Jebel Ali",195,"Arabic","",B+"107144/3a-ka3ba-improv-comedy-show-at-jafza-one-convention-centre"),
("Dubai","Comedy","Father Ted","Dubai Drama Group",date(2026,9,25),date(2026,9,26),"20:00; Sat 15:00 & 20:00","Theatre by QE2, Queen Elizabeth 2, Port Rashid",99,"English","",B+"104715/father-ted-at-theatre-by-qe2"),
("Dubai","Comedy + Desi","Sumukhi Suresh Live in Dubai","Sumukhi Suresh",date(2026,9,26),None,"19:30","Emirates Theatre, Emirates International School, Jumeirah",100,"Mainly English","",B+"106978/sumukhi-suresh-live-in-dubai"),
("Dubai","Desi","KHUSHIYAN: A Live Storytelling Experience","Laksh Maheshwari",date(2026,9,26),None,"20:00","Sheikh Rashid Auditorium, Indian High School, Oud Metha",100,"Hindi","",B+"106836/khushiyan-a-live-storytelling-experience-featuring-laksh-maheshwari"),
("Dubai","Desi","Baithak ft. Osman Mir & Aamir Mir","Osman Mir & Aamir Mir",date(2026,10,3),None,"21:00","Le Meridien Dubai Hotel & Conference Centre, Al Garhoud",157.5,"Not stated","",B+"107705/baithak-ft-osman-mir-and-aamir-mir"),
("Dubai","Comedy","Alaa Abu Diab at The Raj Mahal Theatre","Alaa Abu Diab",date(2026,10,3),None,"21:00","Rajmahal Theatre, Bollywood Parks, Dubai Parks and Resorts",220,"Arabic","",B+"107702/alaa-abu-diab-at-the-raj-mahal-theatre-in-dubai-parks-and-resorts"),
("Dubai","Comedy","Meen Amin by Ahmed Amin Live at The Agenda","Ahmed Amin",date(2026,10,3),None,"20:30","The Agenda, Al Jaddi Street, Dubai Media City",210,"Arabic","",B+"106612/meen-amin-by-ahmed-amin-live-at-the-agenda-dubai"),
("Dubai","Comedy","Shawn Chidiac Live - Laughing in Translation Remix","Shawn Chidiac",date(2026,10,5),None,"20:00","Coca-Cola Arena, City Walk",199,"Not stated","",B+"104860/shawn-chidiac-laughing-in-translation-remix"),
("Dubai","Comedy","Mohandis Al Thauq Al Aam at Dubai Comedy Festival","Al Qaiser Events production",date(2026,10,9),date(2026,10,10),"20:30","Dubai Opera, Downtown Dubai",250,"Arabic","Dubai Comedy Festival",B+"106245/mohandis-al-thauq-al-aam-at-dubai-comedy-festival"),
("Dubai","Comedy","Alaa El Sheikh Stand-up","Alaa El Sheikh",date(2026,10,9),None,"21:00","JAFZA One Convention Centre, Jebel Ali",150,"Arabic","",B+"107443/alaa-el-sheikh-stand-up-at-jafza-one-convention-centre"),
("Dubai","Comedy + Desi","Jamie Lever Live at Dubai Comedy Festival","Jamie Lever",date(2026,10,10),None,"19:00","Emirates Theatre, Emirates International School, Jumeirah",145,"Hindi","Dubai Comedy Festival",B+"107166/jamie-lever-live-in-dubai-comedy-festival"),
("Dubai","Comedy","Munawar Faruqui Live at Dubai Comedy Festival","Munawar Faruqui",date(2026,10,11),None,"18:30","Dubai Opera, Downtown Dubai",125,"Hindi","Dubai Comedy Festival",B+"106537/munawar-faruqui-live-at-dubai-comedy-festival"),
("Dubai","Comedy","Mo Gilligan Live at Dubai Comedy Festival","Mo Gilligan",date(2026,10,12),None,"18:30","Dubai Opera, Downtown Dubai",250,"English","Dubai Comedy Festival",B+"106536/mo-gilligan-live-at-dubai-comedy-festival"),
("Dubai","Comedy","Alexander Merkul Live at Dubai Comedy Festival","Alexander Merkul",date(2026,10,15),date(2026,10,16),"21:30","The New Covent Garden Theatre, Mall of the Emirates",150,"Russian","Dubai Comedy Festival",B+"106503/alexander-merkul-live-at-dubai-comedy-festival"),
("Dubai","Comedy","Zarna Garg Live at Dubai Comedy Festival","Zarna Garg",date(2026,10,16),None,"18:30","Dubai Opera, Downtown Dubai",195,"English","Dubai Comedy Festival",B+"107077/zarna-garg-live-at-dubai-comedy-festival"),
("Dubai","Comedy","Shane Todd Live at Dubai Comedy Festival","Shane Todd",date(2026,10,16),None,"18:30","The New Covent Garden Theatre, Mall of the Emirates",195,"English","Dubai Comedy Festival",B+"106654/shane-todd-live-at-dubai-comedy-festival"),
("Dubai","Comedy + Desi","Abbas Bukhari Live at Dubai Comedy Festival","Abbas Bukhari",date(2026,10,16),None,"21:30","The New Covent Garden Theatre, Mall of the Emirates",125,"Urdu","Dubai Comedy Festival",B+"107641/abbas-bukhari-live-at-dubai-comedy-festival"),
("Dubai","Desi","Navratri Utsav 2026","Live music & dance performances (not named)",date(2026,10,16),date(2026,10,17),"19:30","Zabeel Park Amphitheatre",10,"Not stated","",B+"107326/navratri-utsav-2026-in-dubai"),
("Dubai","Comedy","Game. Set. Match (Re-Run)","Multi-cultural theatrical cast",date(2026,10,16),date(2026,10,18),"Not stated","The Junction, Alserkal Avenue, Al Quoz",105,"English","",B+"105073/gamesetmatch-re-run-at-the-junction-in-dubai"),
("Dubai","Comedy","John Achkar Feena Nehke? Live at Dubai Comedy Festival","John Achkar",date(2026,10,17),None,"18:30","Dubai Opera, Downtown Dubai",185,"Arabic","Dubai Comedy Festival",B+"106524/john-achkar-feena-nehke-live-at-dubai-comedy-festival"),
("Dubai","Comedy","Vir Das Live at Dubai Comedy Festival","Vir Das",date(2026,10,18),None,"20:00","Coca-Cola Arena, City Walk",195,"English","Dubai Comedy Festival",B+"107043/vir-das-live-at-dubai-comedy-festival"),
("Dubai","Comedy + Desi","Amit Tandon Live at Dubai Comedy Festival","Amit Tandon",date(2026,10,18),None,"15:00","Dubai Opera, Downtown Dubai",150,"Hindi","Dubai Comedy Festival",B+"106482/amit-tandon-live-at-dubai-comedy-festival"),
("Dubai","Comedy","Un Air De Famille","La Troupe de Culture Emulsion",date(2026,10,22),date(2026,10,23),"19:45","The Junction, Alserkal Avenue, Al Quoz",160,"French","",B+"107486/un-air-de-famille-at-the-junction-in-dubai"),
("Dubai","Desi","Falguni Pathak Live in Dubai","Falguni Pathak",date(2026,10,24),None,"19:30","Zabeel Park Amphitheatre",60,"Not stated","",B+"107047/falguni-pathak-live-in-dubai"),
("Dubai","Comedy + Desi","Comedy Box Office by Onkar, Shreya & Pranav Sharma","Onkar, Shreya, Pranav Sharma",date(2026,10,25),None,"19:00","Emirates Theatre, Emirates International School, Jumeirah",125,"Not stated","",B+"107402/comedy-box-office-by-onkar-shreya-and-pranav-sharma-in-dubai"),
("Dubai","Comedy + Desi","Kunal Kamra Live in Dubai","Kunal Kamra",date(2026,10,31),None,"18:00","Emirates Theatre, Emirates International School, Jumeirah",155,"Mostly Hindi","",B+"106948/kunal-kamra-live-in-dubai"),
("Dubai","Comedy","Mikhail Shats New Program","Mikhail Shats",date(2026,11,1),None,"20:00","Meyana Theatre, Jumeirah Beach Hotel",550,"Russian","",B+"107195/mikhail-shats-stand-up-comedy-show-at-meyana-theatrejumeirah-beach-hotel-in-dubai"),
("Dubai","Comedy","Olga Malashchenko. Stand-up","Olga Malashchenko",date(2026,11,7),date(2026,11,8),"19:00","Dubai Heights Academy, Al Barsha Road",150,"Russian","",B+"105194/olga-malashchenko-stand-up"),
("Dubai","Comedy","Abdelkader Secteur at Emirates Theatre","Abdelkader Secteur",date(2026,11,8),None,"19:30","Emirates Theatre, Emirates International School, Jumeirah",195,"Arabic","",B+"107374/abdelkader-secteur-in-dubai"),
("Dubai","Comedy","Bassam Wehbe - One Way Ticket","Bassam Wehbe",date(2026,11,19),None,"21:00","Playhouse Studio Theatre, Mall of the Emirates",200,"Arabic with some French","",B+"106609/bassam-wehbe-one-way-ticket"),
("Dubai","Desi","Anuv Jain Live at Expo City Dubai","Anuv Jain",date(2026,11,20),None,"20:00","Jubilee Park, Expo City Dubai",149,"Not stated","",B+"104347/anuv-jain-live-in-expo-city-dubai"),
("Dubai","Comedy + Desi","Out Of Order ft. Shashi Dhiman","Shashi Dhiman",date(2026,11,20),None,"20:45","Live@Play, Warehouse B17, ABA Avenue, Al Quoz",75,"Hindi","",B+"107074/out-of-order-ft-shashi-dhiman-in-dubai"),
("Dubai","Comedy","Trevor Noah Live at Dubai Opera","Trevor Noah",date(2026,11,25),date(2026,11,26),"20:00","Dubai Opera, Downtown Dubai",295,"Not stated","",B+"107317/trevor-noah-live-at-dubai-opera"),
("Dubai","Desi","Atif Aslam 2026 in Dubai","Atif Aslam",date(2026,11,27),None,"21:00","Coca-Cola Arena, City Walk",150,"Not stated","",B+"104859/atif-aslam"),
("Dubai","Desi","URJA - Divine Energy","Sikkil Gurucharan (27th), Kunnakudi Balamuralikrishna (28th), Bharat Sundar (29th)",date(2026,11,27),date(2026,11,29),"18:00","Zabeel Ladies Club (listing title says Oud Mehta Theater)",100,"Carnatic / Indian classical","",B+"104970/urja-divine-energy-at-oud-mehta-theater-in-dubai"),
("Dubai","Comedy + Desi","Comedy Box Office by Jaspreet Singh","Jaspreet Singh",date(2026,11,28),None,"19:00","Emirates Theatre, Emirates International School, Jumeirah",125,"Not stated","",B+"107401/comedy-box-office-by-jaspreet-singh"),
("Dubai","Comedy + Desi","Gaurav Gupta Live in Dubai","Gaurav Gupta",date(2026,12,12),None,"19:00","Sheikh Rashid Auditorium, Indian High School, Oud Metha",95,"Hindi","",B+"gaurav-gupta-live-in-dubai"),
("Dubai","Comedy","Varun Grover Live in Dubai","Varun Grover",date(2027,1,9),None,"18:00","Emirates Theatre, Emirates International School, Jumeirah",125,"Mainly Hindi","",B+"106970/varun-grover-live-in-dubai"),
("Dubai","Comedy + Desi","Biswa Kalyan Rath Live in Dubai 2027","Biswa Kalyan Rath",date(2027,1,23),None,"19:00","Sheikh Rashid Auditorium, Indian High School, Oud Metha",100,"Mainly Hindi","",B+"106982/biswa-kalyan-rath-live-in-dubai"),
("Dubai","Comedy + Desi","Sapan Verma Live in Dubai","Sapan Verma",date(2027,2,6),None,"19:30","Live@Play, Warehouse B17, ABA Avenue, Al Quoz",90,"50% Hindi / 50% English","",B+"106952/sapan-verma-live-in-dubai"),
("Dubai","Comedy + Desi","Silly Papa ft. Angad Ranyal Live in Dubai","Angad Ranyal",date(2027,3,20),None,"19:30","Live@Play by Hive, Al Quoz Industrial Area 2",90,"Mainly Hindi","",B+"107030/silly-papa-ft-angad-ranyal-live-in-dubai"),
("Abu Dhabi","Desi","The Revolution Tour - Sonu Nigam","Sonu Nigam",date(2026,8,21),None,"21:00","Etihad Arena, Yas Island",95,"Not stated","",A+"106429/the-revolution-tour-sonu-nigam"),
("Abu Dhabi","Comedy","Morgan Jay - The Goofy Guy Tour","Morgan Jay",date(2026,10,11),None,"20:00","Etihad Arena, Yas Island",275,"English","",A+"106036/morgan-jay-the-goofy-guy-tour"),
("Abu Dhabi","Comedy + Desi","Russell Peters (Live Nation Middle East)","Russell Peters",date(2026,10,25),None,"20:30","Etihad Arena, Yas Island",275,"Not stated","",A+"104963/live-nation-middle-east-presents-russell-peters-at-etihad-arena"),
("Abu Dhabi","Comedy","MO AMER at Etihad Arena","Mo Amer",date(2026,11,6),None,"20:00","Etihad Arena, Yas Island",325,"Not stated","",A+"107257/mo-amer-at-etihad-arena"),
]

HEAD = ["#","City","Category","Event","Artist / Performer","Start date","End date","Time","Venue","Price from (AED)","Language","Status","Notes","Ticket link"]

wb = Workbook()
ws = wb.active
ws.title = "Events"

FONT = "Arial"
hdr_fill = PatternFill("solid", fgColor="1F3864")
hdr_font = Font(name=FONT, bold=True, color="FFFFFF", size=11)
thin = Side(style="thin", color="D0D0D0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

ws["A1"] = "Comedy & Desi events - Dubai and Abu Dhabi (Platinumlist)"
ws["A1"].font = Font(name=FONT, bold=True, size=14)
ws["A2"] = "Compiled 13 Aug 2026 from Platinumlist comedy and desi listings; each row verified on its own event page. Blank end date = single-day event. 'Status' updates automatically against today's date."
ws["A2"].font = Font(name=FONT, italic=True, size=9, color="595959")
ws.merge_cells("A1:N1")
ws.merge_cells("A2:N2")

HR = 4
for c, h in enumerate(HEAD, 1):
    cell = ws.cell(row=HR, column=c, value=h)
    cell.fill = hdr_fill
    cell.font = hdr_font
    cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    cell.border = border

city_fill = {"Dubai": PatternFill("solid", fgColor="FFFFFF"), "Abu Dhabi": PatternFill("solid", fgColor="FFF7E6")}

r = HR + 1
for i, (city, cat, ev, art, sd, ed, tm, ven, price, lang, notes, url) in enumerate(rows, 1):
    ws.cell(row=r, column=1, value=i)
    ws.cell(row=r, column=2, value=city)
    ws.cell(row=r, column=3, value=cat)
    ws.cell(row=r, column=4, value=ev)
    ws.cell(row=r, column=5, value=art)
    ws.cell(row=r, column=6, value=sd).number_format = "ddd d mmm yyyy"
    c7 = ws.cell(row=r, column=7, value=ed)
    c7.number_format = "ddd d mmm yyyy"
    ws.cell(row=r, column=8, value=tm)
    ws.cell(row=r, column=9, value=ven)
    ws.cell(row=r, column=10, value=price).number_format = '#,##0.00'
    ws.cell(row=r, column=11, value=lang)
    ws.cell(row=r, column=12, value=f'=IF(IF(G{r}="",F{r},G{r})<TODAY(),"Past","Upcoming")')
    ws.cell(row=r, column=13, value=notes)
    link = ws.cell(row=r, column=14, value="Book tickets")
    link.hyperlink = url
    link.font = Font(name=FONT, color="0563C1", underline="single", size=10)
    for c in range(1, 15):
        cl = ws.cell(row=r, column=c)
        if c != 14:
            cl.font = Font(name=FONT, size=10)
        cl.border = border
        cl.alignment = Alignment(vertical="top", wrap_text=(c in (4, 5, 9, 13)))
        cl.fill = city_fill[city]
    r += 1

last = r - 1
widths = {"A":5,"B":11,"C":15,"D":42,"E":34,"F":16,"G":16,"H":18,"I":42,"J":15,"K":22,"L":11,"M":40,"N":14}
for col, w in widths.items():
    ws.column_dimensions[col].width = w
ws.row_dimensions[HR].height = 32
ws.freeze_panes = f"D{HR+1}"
ws.auto_filter.ref = f"A{HR}:N{last}"

# Summary sheet
s = wb.create_sheet("Summary")
s["A1"] = "Summary"
s["A1"].font = Font(name=FONT, bold=True, size=14)
labels = [
    ("Total events", f'=COUNTA(Events!D{HR+1}:D{last})'),
    ("Dubai", f'=COUNTIF(Events!B{HR+1}:B{last},"Dubai")'),
    ("Abu Dhabi", f'=COUNTIF(Events!B{HR+1}:B{last},"Abu Dhabi")'),
    ("Comedy only", f'=COUNTIF(Events!C{HR+1}:C{last},"Comedy")'),
    ("Desi only", f'=COUNTIF(Events!C{HR+1}:C{last},"Desi")'),
    ("Listed in both categories", f'=COUNTIF(Events!C{HR+1}:C{last},"Comedy + Desi")'),
    ("Still upcoming", f'=COUNTIF(Events!L{HR+1}:L{last},"Upcoming")'),
    ("Dubai Comedy Festival shows", f'=COUNTIF(Events!M{HR+1}:M{last},"Dubai Comedy Festival")'),
    ("Cheapest ticket (AED)", f'=MIN(Events!J{HR+1}:J{last})'),
    ("Dearest ticket (AED)", f'=MAX(Events!J{HR+1}:J{last})'),
    ("Median ticket (AED)", f'=MEDIAN(Events!J{HR+1}:J{last})'),
]
rr = 3
for lab, f in labels:
    s.cell(row=rr, column=1, value=lab).font = Font(name=FONT, size=10, bold=True)
    s.cell(row=rr, column=2, value=f).font = Font(name=FONT, size=10)
    rr += 1
s.column_dimensions["A"].width = 30
s.column_dimensions["B"].width = 14
s.cell(row=rr+1, column=1, value="Source: Platinumlist Dubai (dubai.platinumlist.net) and Abu Dhabi (abu-dhabi.platinumlist.net), comedy and desi category listings.").font = Font(name=FONT, italic=True, size=9, color="595959")

wb.save("/home/claude/out/UAE_Comedy_Desi_Events.xlsx")
print("rows:", len(rows))
