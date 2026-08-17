"""
Import an event checklist workbook -> data/checklists.json

The source workbooks have three sheets: Setup (show name, key dates and assumptions),
Dashboard (formulas only, recomputed by the site so nothing is read from it) and
Checklist (the tasks themselves).

Task due dates are not stored. Each task carries a D-minus offset from the show date,
exactly as the workbook does, so setting the show date on the site recalculates every
due date. A negative offset falls after the show.

Run:  python src/import_checklist.py <workbook.xlsx> [--id my-show]
"""
import argparse
import json
import re
import sys
import unicodedata
from datetime import date, datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:                                          # pragma: no cover
    print("openpyxl is needed to import a workbook: pip install openpyxl")
    raise SystemExit(1)

ROOT = Path(__file__).resolve().parent.parent
STATUSES = ["Not started", "In progress", "Done", "Not needed"]

# Header labels in the Checklist sheet mapped to the keys used in checklists.json.
COLUMNS = {
    "#": "n",
    "workstream": "workstream",
    "task": "task",
    "why it matters": "why",
    "owner": "owner",
    "blocking": "blocking",
    "d-minus": "d_minus",
    "status": "status",
}


def clean(value):
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.date().isoformat() if isinstance(value, datetime) else value.isoformat()
    text = " ".join(str(value).split())
    # The workbooks use em and en dashes; the site's copy uses plain hyphens.
    return text.replace("—", " - ").replace("–", "-").strip()


def slugify(text):
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode()
    return re.sub(r"-+", "-", re.sub(r"[^a-z0-9]+", "-", text.lower())).strip("-")


def title_case(text):
    """Workbook headings are shouted; the site is not."""
    if text and text == text.upper():
        return " - ".join(part.strip().title() for part in text.split(" - "))
    return text


def read_setup(ws):
    """Rows are label / value / note, with a title and subtitle above them."""
    title = subtitle = ""
    fields, seen_key_dates = [], False
    for row in ws.iter_rows(values_only=True):
        cells = [clean(c) for c in row]
        text = [c for c in cells if c]
        if not text:
            continue
        joined = " ".join(text)
        if not title:
            title = title_case(joined)
            continue
        if not subtitle and "KEY DATES" not in joined.upper():
            subtitle = joined.rstrip(" ·")
            continue
        if "KEY DATES" in joined.upper():
            seen_key_dates = True
            continue
        if "HOW TO USE" in joined.upper():
            break
        if seen_key_dates and len(cells) >= 2 and cells[1]:
            label = cells[1]
            if label.lower() == "today":          # recomputed at view time
                continue
            fields.append({"label": label,
                           "value": cells[2] if len(cells) > 2 else "",
                           "note": cells[3] if len(cells) > 3 else ""})
    return title, subtitle, fields


def read_tasks(ws):
    # Track the header by row number. values_only=True yields fresh tuples on every
    # pass, so an identity comparison against a row from an earlier pass never matches.
    header_at, index = None, {}
    for number, row in enumerate(ws.iter_rows(values_only=True), 1):
        cells = [clean(c).lower() for c in row]
        if "workstream" in cells and "task" in cells:
            header_at = number
            for pos, name in enumerate(cells):
                if name in COLUMNS:
                    index[COLUMNS[name]] = pos
            break
    if header_at is None:
        raise SystemExit("could not find the Checklist header row")

    tasks = []
    for number, row in enumerate(ws.iter_rows(values_only=True), 1):
        if number <= header_at:
            continue
        cells = [clean(c) for c in row]
        get = lambda key: cells[index[key]] if key in index and index[key] < len(cells) else ""
        if not get("task"):
            continue
        status = get("status") or STATUSES[0]
        d_minus = get("d_minus")
        tasks.append({
            "n": int(get("n")) if get("n").lstrip("-").isdigit() else len(tasks) + 1,
            "workstream": get("workstream"),
            "task": get("task"),
            "why": get("why"),
            "owner": get("owner"),
            "blocking": get("blocking").strip().lower() in ("yes", "true", "y"),
            "d_minus": int(d_minus) if d_minus.lstrip("-").isdigit() else None,
            "status": status if status in STATUSES else STATUSES[0],
        })
    return tasks


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("workbook")
    ap.add_argument("--id", help="stable id; defaults to a slug of the title")
    ap.add_argument("--out", default=str(ROOT / "data" / "checklists.json"))
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.workbook, data_only=True)
    setup = wb["Setup"] if "Setup" in wb.sheetnames else wb.worksheets[0]
    sheet = wb["Checklist"] if "Checklist" in wb.sheetnames else wb.worksheets[-1]

    title, subtitle, fields = read_setup(setup)
    tasks = read_tasks(sheet)
    checklist = {
        "id": args.id or slugify(title),
        "title": title,
        "subtitle": subtitle,
        "show_date": None,
        "setup": fields,
        "tasks": tasks,
    }

    out = Path(args.out)
    existing = {"checklists": []}
    if out.exists():
        existing = json.loads(out.read_text())
    # Replace in place if the id is already present, so a re-import is idempotent.
    others = [c for c in existing.get("checklists", []) if c["id"] != checklist["id"]]
    out.write_text(json.dumps({"checklists": others + [checklist]}, indent=1,
                              ensure_ascii=False) + "\n")

    streams = {}
    for t in tasks:
        streams[t["workstream"]] = streams.get(t["workstream"], 0) + 1
    print(f"imported '{checklist['title']}' as {checklist['id']}")
    print(f"  {len(tasks)} tasks, {sum(1 for t in tasks if t['blocking'])} blocking")
    print(f"  workstreams: {', '.join(f'{k} {v}' for k, v in streams.items())}")
    print(f"  wrote {out}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
